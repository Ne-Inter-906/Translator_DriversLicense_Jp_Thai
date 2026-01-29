from openpyxl import load_workbook
import openpyxl
from copy import copy

class Layout_Manager:
    def __init__(self):
        pass

    def sync_layout(self, template_path, target_path, output_path):
        print(f"レイアウト調整を開始: {target_path}")
        
        wb_temp = openpyxl.load_workbook(template_path)
        wb_target = openpyxl.load_workbook(target_path)

        ws_temp = wb_temp.worksheets[0] 
        ws_target = wb_target.worksheets[0]

        # 1. セル結合の同期
        print("セルの結合を同期中...")
        if ws_target.merged_cells:
            for merged_range in list(ws_target.merged_cells.ranges):
                ws_target.unmerge_cells(str(merged_range))
        
        for merged_range in ws_temp.merged_cells.ranges:
            ws_target.merge_cells(str(merged_range))

        # 2. 列の幅は同期し、行の高さはリセットする
        print("列の幅を同期中...")
        for key, dim in ws_temp.column_dimensions.items():
            ws_target.column_dimensions[key].width = dim.width
            
        print("行の高さを自動調整用にリセット中...")
        for row_idx in range(1, ws_target.max_row + 1):
            # 高さをNoneに設定するだけで、Excelは自動的に「カスタムではない」と判断します
            ws_target.row_dimensions[row_idx].height = None
            # ws_target.row_dimensions[row_idx].customHeight = False  <-- この行を削除

            
        # 3. 全セルのスタイルを詳細コピー
        print("全セルのスタイルを詳細コピー中...")
        for row in ws_temp.iter_rows():
            for cell_temp in row:
                cell_target = ws_target.cell(row=cell_temp.row, column=cell_temp.column)
                if cell_temp.has_style:
                    cell_target.font = copy(cell_temp.font)
                    cell_target.border = copy(cell_temp.border)
                    cell_target.fill = copy(cell_temp.fill)
                    
                    # alignmentは一括コピーしてからwrapTextを上書き
                    new_alignment = copy(cell_temp.alignment)
                    new_alignment.wrapText = True 
                    cell_target.alignment = new_alignment

        wb_target.save(output_path)
        print(f"レイアウト調整が完了しました: {output_path}")