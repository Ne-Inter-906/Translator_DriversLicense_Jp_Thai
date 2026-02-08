import pandas as pd
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

class SimilarityChecker:
    def __init__(self, model_name='sentence-transformers/LaBSE'):
        print("LaBSEモデルの読み込み中...")
        self.model = SentenceTransformer(model_name)
        self.threshold = 0.75
    
    def check_file(self, file_path, target_columns):
        if not os.path.exists(file_path):
            print(f"ファイルが見つかりません: {file_path}")
            return

        print(f"チェック開始: {file_path}")

        wb = load_workbook(file_path)
        ws = wb.worksheets[0] 

        header = [cell.value for cell in ws[1]]
        try:
            # 各列のインデックスを取得
            idx_org_Q = header.index(target_columns['問題文']) + 1
            idx_tr_Q = header.index(target_columns['ปัญหา']) + 1
            idx_org_E = header.index(target_columns['解説']) + 1
            idx_tr_E = header.index(target_columns['คำอธิบาย']) + 1
        except ValueError as e:
            print(f"error: 指定された列名が見つかりません。 {e}")
            return
        
        # 新しい列を末尾に追加
        score_col_Q = len(header) + 1
        score_col_E = len(header) + 2
        ws.cell(row=1, column=score_col_Q).value = "問題文_翻訳精度スコア"
        ws.cell(row=1, column=score_col_E).value = "解説_翻訳精度スコア"

        red_font = Font(color="FF0000", bold=True)

        for r in range(2, ws.max_row + 1):
            # --- 問題文(Question)のチェック ---
            val_org_Q = ws.cell(row=r, column=idx_org_Q).value
            val_tr_Q = ws.cell(row=r, column=idx_tr_Q).value
            
            if val_org_Q and val_tr_Q:
                score_q = self._get_similarity(val_org_Q, val_tr_Q)
                cell_score_q = ws.cell(row=r, column=score_col_Q)
                cell_score_q.value = score_q
                
                if score_q < self.threshold:
                    ws.cell(row=r, column=idx_tr_Q).font = red_font
                    cell_score_q.font = red_font

            # --- 解説(Explanation)のチェック ---
            val_org_E = ws.cell(row=r, column=idx_org_E).value
            val_tr_E = ws.cell(row=r, column=idx_tr_E).value
            
            if val_org_E and val_tr_E:
                score_e = self._get_similarity(val_org_E, val_tr_E)
                cell_score_e = ws.cell(row=r, column=score_col_E)
                cell_score_e.value = score_e
                
                if score_e < self.threshold:
                    ws.cell(row=r, column=idx_tr_E).font = red_font
                    cell_score_e.font = red_font

        output_path = file_path.replace(".xlsx", "_checked.xlsx")
        wb.save(output_path)
        print(f"チェックが完了しました: {output_path}")
        return output_path

    def _get_similarity(self, text1, text2):
        """2つのテキストの類似度を計算する内部メソッド"""
        emb1 = self.model.encode(str(text1))
        emb2 = self.model.encode(str(text2))
        return round(util.cos_sim(emb1, emb2).item(), 4)